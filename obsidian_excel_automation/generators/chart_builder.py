"""
Chart builder for Excel workbooks
Creates bar charts, pie charts, and line charts
"""

from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional, List


class ChartBuilder:
    """Builds charts for Excel workbooks"""

    @staticmethod
    def add_bar_chart(
        ws: Worksheet,
        title: str,
        data_range: str,
        categories_range: Optional[str] = None,
        position: str = "A1",
        width: int = 15,
        height: int = 10,
        horizontal: bool = False
    ):
        """
        Add a bar chart to worksheet

        Args:
            ws: Worksheet
            title: Chart title
            data_range: Data range (e.g., "B2:B10")
            categories_range: Categories range (e.g., "A2:A10")
            position: Chart position (cell reference)
            width: Chart width in characters
            height: Chart height in characters
            horizontal: True for horizontal bars
        """
        chart = BarChart()
        chart.type = "bar" if horizontal else "col"
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height

        # Data
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=False)

        # Categories
        if categories_range:
            cats = Reference(ws, range_string=categories_range)
            chart.set_categories(cats)

        # Add to worksheet
        ws.add_chart(chart, position)

    @staticmethod
    def add_pie_chart(
        ws: Worksheet,
        title: str,
        data_range: str,
        categories_range: Optional[str] = None,
        position: str = "A1",
        width: int = 12,
        height: int = 10,
        show_percentage: bool = True
    ):
        """
        Add a pie chart to worksheet

        Args:
            ws: Worksheet
            title: Chart title
            data_range: Data range (e.g., "B2:B10")
            categories_range: Categories range (e.g., "A2:A10")
            position: Chart position (cell reference)
            width: Chart width in characters
            height: Chart height in characters
            show_percentage: Show percentage labels
        """
        chart = PieChart()
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height

        # Data
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=False)

        # Categories
        if categories_range:
            cats = Reference(ws, range_string=categories_range)
            chart.set_categories(cats)

        # Data labels
        if show_percentage:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showCatName = True

        # Add to worksheet
        ws.add_chart(chart, position)

    @staticmethod
    def add_line_chart(
        ws: Worksheet,
        title: str,
        data_range: str,
        categories_range: Optional[str] = None,
        position: str = "A1",
        width: int = 15,
        height: int = 10,
        smooth: bool = True
    ):
        """
        Add a line chart to worksheet

        Args:
            ws: Worksheet
            title: Chart title
            data_range: Data range (e.g., "B2:B10")
            categories_range: Categories range (e.g., "A2:A10")
            position: Chart position (cell reference)
            width: Chart width in characters
            height: Chart height in characters
            smooth: Use smooth lines
        """
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height
        chart.smooth = smooth

        # Data
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=False)

        # Categories
        if categories_range:
            cats = Reference(ws, range_string=categories_range)
            chart.set_categories(cats)

        # Add to worksheet
        ws.add_chart(chart, position)

    @staticmethod
    def add_progress_chart(
        ws: Worksheet,
        title: str,
        categories: List[str],
        completed: List[int],
        total: List[int],
        position: str = "A1",
        width: int = 15,
        height: int = 10
    ):
        """
        Add a stacked bar chart showing progress (completed vs remaining)

        Args:
            ws: Worksheet
            title: Chart title
            categories: Category labels
            completed: Completed values
            total: Total values
            position: Chart position
            width: Chart width
            height: Chart height
        """
        # Write temporary data for chart
        start_row = ws.max_row + 2
        ws.cell(row=start_row, column=1).value = "Category"
        ws.cell(row=start_row, column=2).value = "Completed"
        ws.cell(row=start_row, column=3).value = "Remaining"

        for idx, (cat, comp, tot) in enumerate(zip(categories, completed, total), start=1):
            ws.cell(row=start_row + idx, column=1).value = cat
            ws.cell(row=start_row + idx, column=2).value = comp
            ws.cell(row=start_row + idx, column=3).value = tot - comp

        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height

        # Data
        data = Reference(
            ws,
            min_col=2,
            min_row=start_row,
            max_col=3,
            max_row=start_row + len(categories)
        )
        cats = Reference(
            ws,
            min_col=1,
            min_row=start_row + 1,
            max_row=start_row + len(categories)
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Add to worksheet
        ws.add_chart(chart, position)
