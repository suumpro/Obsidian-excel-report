"""
Base Excel generator - creates and manages Excel workbooks
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from typing import List, Dict, Any, Optional

from .style_manager import StyleManager
from ..config import Config
from ..utils.logger import logger


class ExcelGenerator:
    """Base class for generating Excel workbooks"""

    def __init__(self, config: Config):
        """
        Initialize Excel generator

        Args:
            config: Application configuration
        """
        self.config = config
        self.style_manager = StyleManager(config.styling)
        self.workbook: Optional[Workbook] = None

    def create_workbook(self) -> Workbook:
        """Create a new Excel workbook"""
        logger.info("Creating new Excel workbook...")
        self.workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

        return self.workbook

    def add_sheet(self, title: str, index: Optional[int] = None) -> Worksheet:
        """
        Add a new sheet to the workbook

        Args:
            title: Sheet title
            index: Position (None = append at end)

        Returns:
            Created worksheet
        """
        if not self.workbook:
            self.create_workbook()

        logger.debug(f"Adding sheet: {title}")

        if index is not None:
            ws = self.workbook.create_sheet(title, index)
        else:
            ws = self.workbook.create_sheet(title)

        return ws

    def save(self, output_path: Path):
        """
        Save workbook to file

        Args:
            output_path: Output file path
        """
        if not self.workbook:
            raise ValueError("No workbook to save")

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Saving workbook to: {output_path}")
        self.workbook.save(output_path)
        logger.info("✅ Workbook saved successfully")

    def add_summary_section(
        self,
        ws: Worksheet,
        title: str,
        data: Dict[str, Any],
        start_row: int,
        start_col: int = 1
    ) -> int:
        """
        Add a summary section with key-value pairs

        Args:
            ws: Worksheet
            title: Section title
            data: Dictionary of key-value pairs
            start_row: Starting row
            start_col: Starting column

        Returns:
            Next available row
        """
        current_row = start_row

        # Title
        cell = ws.cell(row=current_row, column=start_col)
        cell.value = title
        self.style_manager.apply_subheader_style(cell)
        ws.merge_cells(
            start_row=current_row,
            start_column=start_col,
            end_row=current_row,
            end_column=start_col + 1
        )
        current_row += 1

        # Data rows
        for key, value in data.items():
            # Key
            key_cell = ws.cell(row=current_row, column=start_col)
            self.style_manager.apply_normal_style(key_cell, key, bold=True)

            # Value
            value_cell = ws.cell(row=current_row, column=start_col + 1)
            self.style_manager.apply_normal_style(value_cell, str(value))

            current_row += 1

        return current_row + 1  # Add spacing

    def add_table(
        self,
        ws: Worksheet,
        headers: List[str],
        data: List[List[Any]],
        start_row: int,
        start_col: int = 1,
        apply_alternating_colors: bool = True
    ) -> int:
        """
        Add a table with headers and data

        Args:
            ws: Worksheet
            headers: Column headers
            data: Data rows
            start_row: Starting row
            start_col: Starting column
            apply_alternating_colors: Apply alternating row colors

        Returns:
            Next available row
        """
        current_row = start_row

        # Headers
        self.style_manager.format_header_row(ws, current_row, headers, start_col)
        current_row += 1

        # Data
        for row_data in data:
            self.style_manager.format_data_row(ws, current_row, row_data, start_col)
            current_row += 1

        # Apply table styling
        if apply_alternating_colors and data:
            end_row = start_row + len(data)
            end_col = start_col + len(headers) - 1
            self.style_manager.apply_table_style(
                ws, start_row + 1, end_row, start_col, end_col
            )

        return current_row + 1  # Add spacing

    def add_metric_boxes(
        self,
        ws: Worksheet,
        metrics: List[Dict[str, Any]],
        start_row: int,
        start_col: int = 1,
        boxes_per_row: int = 4
    ) -> int:
        """
        Add metric boxes in a grid layout

        Args:
            ws: Worksheet
            metrics: List of {title, value, target} dicts
            start_row: Starting row
            start_col: Starting column
            boxes_per_row: Number of boxes per row

        Returns:
            Next available row
        """
        current_row = start_row
        current_col = start_col

        for idx, metric in enumerate(metrics):
            # Box title
            title_cell = ws.cell(row=current_row, column=current_col)
            title_cell.value = metric.get('title', '')
            self.style_manager.apply_subheader_style(title_cell)

            # Merge cells for box
            ws.merge_cells(
                start_row=current_row,
                start_column=current_col,
                end_row=current_row,
                end_column=current_col + 1
            )

            # Value row
            value_row = current_row + 1

            # Current value
            value_cell = ws.cell(row=value_row, column=current_col)
            value_cell.value = metric.get('value', '')
            self.style_manager.apply_normal_style(value_cell, bold=True, centered=True)

            # Target value (if provided)
            if 'target' in metric:
                target_cell = ws.cell(row=value_row, column=current_col + 1)
                target_cell.value = f"목표: {metric['target']}"
                self.style_manager.apply_normal_style(target_cell, centered=True)

            # Progress row (if provided)
            if 'progress' in metric:
                progress_row = current_row + 2
                progress_cell = ws.cell(row=progress_row, column=current_col)
                progress_cell.value = metric['progress']
                self.style_manager.apply_normal_style(progress_cell, centered=True)
                ws.merge_cells(
                    start_row=progress_row,
                    start_column=current_col,
                    end_row=progress_row,
                    end_column=current_col + 1
                )

            # Move to next position
            if (idx + 1) % boxes_per_row == 0:
                # New row
                current_row += 4  # Box height (title + value + progress + spacing)
                current_col = start_col
            else:
                # Next column
                current_col += 3  # Box width + spacing

        return current_row + 4 if current_col > start_col else current_row

    def set_sheet_properties(
        self,
        ws: Worksheet,
        freeze_panes: bool = True,
        auto_filter: bool = True,
        column_widths: Optional[Dict[int, int]] = None
    ):
        """
        Set common sheet properties

        Args:
            ws: Worksheet
            freeze_panes: Freeze top row
            auto_filter: Enable auto filter on headers
            column_widths: Custom column widths
        """
        if freeze_panes:
            self.style_manager.freeze_panes(ws, row=2, col=1)

        if auto_filter and ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        if column_widths:
            self.style_manager.set_column_widths(ws, column_widths)
        else:
            self.style_manager.auto_adjust_column_widths(ws)
