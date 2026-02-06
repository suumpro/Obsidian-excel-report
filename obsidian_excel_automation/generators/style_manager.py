"""
Style manager for Excel workbooks
Handles colors, fonts, borders, and conditional formatting
"""

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Any, Optional


class StyleManager:
    """Manages Excel styling"""

    def __init__(self, styling_config: Dict[str, Any]):
        """
        Initialize with styling configuration

        Args:
            styling_config: Styling section from config
        """
        self.config = styling_config

        # Define standard styles
        self.header_fill = PatternFill(
            start_color=styling_config.get('header_color', '4472C4'),
            end_color=styling_config.get('header_color', '4472C4'),
            fill_type="solid"
        )

        self.subheader_fill = PatternFill(
            start_color=styling_config.get('subheader_color', 'D9E1F2'),
            end_color=styling_config.get('subheader_color', 'D9E1F2'),
            fill_type="solid"
        )

        # Priority colors
        priority_colors = styling_config.get('priority_colors', {})
        self.p0_fill = PatternFill(
            start_color=priority_colors.get('P0', 'FFC7CE'),
            end_color=priority_colors.get('P0', 'FFC7CE'),
            fill_type="solid"
        )
        self.p1_fill = PatternFill(
            start_color=priority_colors.get('P1', 'FFEB9C'),
            end_color=priority_colors.get('P1', 'FFEB9C'),
            fill_type="solid"
        )
        self.p2_fill = PatternFill(
            start_color=priority_colors.get('P2', 'C6EFCE'),
            end_color=priority_colors.get('P2', 'C6EFCE'),
            fill_type="solid"
        )

        # Status colors
        status_colors = styling_config.get('status_colors', {})
        self.completed_fill = PatternFill(
            start_color=status_colors.get('completed', 'C6EFCE'),
            end_color=status_colors.get('completed', 'C6EFCE'),
            fill_type="solid"
        )
        self.in_progress_fill = PatternFill(
            start_color=status_colors.get('in_progress', 'FFEB9C'),
            end_color=status_colors.get('in_progress', 'FFEB9C'),
            fill_type="solid"
        )
        self.pending_fill = PatternFill(
            start_color=status_colors.get('pending', 'FFC7CE'),
            end_color=status_colors.get('pending', 'FFC7CE'),
            fill_type="solid"
        )

        # Standard fonts
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.subheader_font = Font(bold=True, size=11)
        self.normal_font = Font(size=10)
        self.bold_font = Font(bold=True, size=10)

        # Alignments
        self.center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.right_alignment = Alignment(horizontal="right", vertical="center")

        # Borders
        thin_side = Side(style="thin", color="000000")
        self.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def apply_header_style(self, cell, text: str = None):
        """Apply header style to a cell"""
        if text:
            cell.value = text
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_alignment
        cell.border = self.border

    def apply_subheader_style(self, cell, text: str = None):
        """Apply subheader style to a cell"""
        if text:
            cell.value = text
        cell.font = self.subheader_font
        cell.fill = self.subheader_fill
        cell.alignment = self.center_alignment
        cell.border = self.border

    def apply_normal_style(self, cell, text: str = None, bold: bool = False, centered: bool = False):
        """Apply normal cell style"""
        if text is not None:
            cell.value = text
        cell.font = self.bold_font if bold else self.normal_font
        cell.alignment = self.center_alignment if centered else self.left_alignment
        cell.border = self.border

    def apply_priority_style(self, cell, priority: str):
        """Apply priority-based coloring"""
        if priority == 'P0':
            cell.fill = self.p0_fill
        elif priority == 'P1':
            cell.fill = self.p1_fill
        elif priority == 'P2':
            cell.fill = self.p2_fill
        cell.border = self.border

    def apply_status_style(self, cell, status: str):
        """Apply status-based coloring"""
        status_lower = status.lower()
        if '완료' in status or 'completed' in status_lower or '✅' in status:
            cell.fill = self.completed_fill
        elif '진행' in status or 'progress' in status_lower or '🔄' in status:
            cell.fill = self.in_progress_fill
        elif '대기' in status or 'pending' in status_lower or '⚠️' in status:
            cell.fill = self.pending_fill
        cell.border = self.border

    def format_header_row(self, ws: Worksheet, row: int, headers: list, start_col: int = 1):
        """Format an entire header row"""
        for col_idx, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=row, column=col_idx)
            self.apply_header_style(cell, header)

    def format_data_row(self, ws: Worksheet, row: int, data: list, start_col: int = 1):
        """Format a data row with normal style"""
        for col_idx, value in enumerate(data, start=start_col):
            cell = ws.cell(row=row, column=col_idx)
            self.apply_normal_style(cell, str(value) if value is not None else '')

    def set_column_widths(self, ws: Worksheet, widths: Dict[int, int]):
        """
        Set column widths

        Args:
            ws: Worksheet
            widths: {column_number: width_in_characters}
        """
        from openpyxl.utils import get_column_letter

        for col_num, width in widths.items():
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width

    def auto_adjust_column_widths(self, ws: Worksheet, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content"""
        from openpyxl.utils import get_column_letter

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    def freeze_panes(self, ws: Worksheet, row: int = 2, col: int = 1):
        """
        Freeze panes at specified position

        Args:
            ws: Worksheet
            row: Row number to freeze above
            col: Column number to freeze left of
        """
        from openpyxl.utils import get_column_letter

        cell_ref = f"{get_column_letter(col)}{row}"
        ws.freeze_panes = cell_ref

    def apply_table_style(self, ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int):
        """Apply alternating row colors to a table"""
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for row_idx in range(start_row, end_row + 1):
            if row_idx % 2 == 0:  # Even rows
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if not cell.fill or cell.fill.start_color.rgb == '00000000':  # No existing fill
                        cell.fill = light_fill
                    cell.border = self.border
