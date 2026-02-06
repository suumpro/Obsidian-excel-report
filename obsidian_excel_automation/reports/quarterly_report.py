"""
Quarterly Report Generator
Generates 4-sheet quarterly status report with KPIs, tasks, and progress charts
"""

from typing import Optional
from datetime import datetime
from pathlib import Path

from ..generators.excel_generator import ExcelGenerator
from ..aggregators.data_aggregator import DataAggregator
from ..utils.date_utils import get_current_quarter_info
from ..utils.logger import logger


class QuarterlyReportGenerator(ExcelGenerator):
    """
    Generates quarterly status report with 4 sheets

    Sheets:
        1. Q{n} Overview - Quarter summary with KPIs
        2. P0 Tasks - Critical tasks breakdown
        3. P1 Tasks - High priority tasks breakdown
        4. Progress Charts - Visual analytics (placeholder for now)
    """

    def __init__(self, config):
        """
        Initialize quarterly report generator

        Args:
            config: Config object with vault paths and settings
        """
        super().__init__(config)
        self.aggregator = DataAggregator(config)

    def generate(
        self,
        quarter: Optional[int] = None,
        year: Optional[int] = None,
        output_filename: Optional[str] = None,
        date: Optional[datetime] = None
    ) -> Path:
        """
        Generate quarterly status report

        Args:
            quarter: Quarter number (1-4), auto-detected if None
            year: Year (defaults to current year)
            output_filename: Custom filename (optional)
            date: Report date (defaults to today)

        Returns:
            Path to generated Excel file

        Raises:
            FileNotFoundError: If quarterly status file not found
            ValueError: If quarter is invalid
        """
        # Auto-detect quarter and year if not provided
        if quarter is None or year is None:
            quarter_info = get_current_quarter_info()
            if quarter is None:
                quarter = quarter_info['quarter']
            if year is None:
                year = quarter_info['year']

        # Validate quarter
        if quarter not in [1, 2, 3, 4]:
            raise ValueError(f"Invalid quarter: {quarter}. Must be 1-4.")

        logger.info(f"🚀 Generating Q{quarter} {year} Status Report...")

        # Load data
        logger.info(f"Loading Q{quarter} data from Obsidian files...")
        try:
            quarterly_data = self.aggregator.load_quarterly_data(quarter)
            dashboard_data = self.aggregator.load_dashboard_data()
        except FileNotFoundError as e:
            logger.error(f"Could not find quarterly data file: {e}")
            logger.error(f"Expected: 00_Dashboard/2026_Q{quarter}_Status.md")
            raise

        # Calculate metrics
        from ..aggregators.metrics_calculator import MetricsCalculator
        metrics = MetricsCalculator.calculate_quarterly_metrics(quarterly_data)

        # Create workbook
        self.wb = self.create_workbook()

        # Build quarter info for display
        current_quarter_info = get_current_quarter_info()
        is_current_quarter = (quarter == current_quarter_info['quarter'] and year == current_quarter_info['year'])

        if is_current_quarter:
            quarter_display_info = current_quarter_info
        else:
            quarter_display_info = {
                'quarter': quarter,
                'year': year,
                'start_date': f"{year}-{(quarter-1)*3+1:02d}-01",
                'end_date': f"{year}-{quarter*3:02d}-{[31,30,31,31][quarter-1]}"
            }

        # Create 4 sheets
        logger.info("Creating Sheet 1: Q{} Overview...".format(quarter))
        self._create_sheet1_overview(quarterly_data, metrics, quarter_display_info)

        logger.info("Creating Sheet 2: P0 Tasks...")
        self._create_sheet2_p0_tasks(quarterly_data)

        logger.info("Creating Sheet 3: P1 Tasks...")
        self._create_sheet3_p1_tasks(quarterly_data)

        logger.info("Creating Sheet 4: Progress Charts...")
        self._create_sheet4_progress_charts(quarterly_data, metrics)

        # Generate filename
        if not output_filename:
            date_obj = date or datetime.now()
            date_str = date_obj.strftime('%Y%m%d')
            output_filename = f"STOREAGENT_Q{quarter}_Status_{date_str}.xlsx"

        # Save
        output_path = self.config.get_output_path(output_filename)
        self.save(output_path)

        logger.info(f"✅ Workbook saved successfully")
        return output_path

    def _create_sheet1_overview(self, quarterly_data, metrics, quarter_info):
        """
        Sheet 1: Q{n} Overview

        Layout:
            - Title and metadata
            - KPI metrics boxes (4x1 grid)
            - Quarter timeline
            - Key achievements
            - Risks and issues
        """
        ws = self.add_sheet(f"Q{quarter_info['quarter']} Overview", 0)

        # Title
        quarter = quarter_info['quarter']
        year = quarter_info['year']
        ws['A1'] = f"Q{quarter} {year} Status Overview"
        ws.merge_cells('A1:F1')
        self.style_manager.apply_header_style(ws['A1'])
        ws.row_dimensions[1].height = 30

        # Metadata section
        row = 3
        ws[f'A{row}'] = "Quarter:"
        ws[f'B{row}'] = f"Q{quarter}"
        ws[f'D{row}'] = "Period:"
        ws[f'E{row}'] = f"{quarter_info.get('start_date', 'N/A')} ~ {quarter_info.get('end_date', 'N/A')}"

        row += 1
        ws[f'A{row}'] = "Report Date:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d')

        # KPI Metrics boxes
        row += 2
        ws[f'A{row}'] = "Key Performance Indicators"
        ws.merge_cells(f'A{row}:F{row}')
        self.style_manager.apply_subheader_style(ws[f'A{row}'])

        row += 1
        kpi_metrics = [
            {
                'label': 'Total Tasks',
                'value': str(metrics.total_tasks),
                'color': None
            },
            {
                'label': 'P0 Completion',
                'value': f"{metrics.p0_completion_rate:.1f}%",
                'color': 'p0'
            },
            {
                'label': 'P1 Completion',
                'value': f"{metrics.p1_completion_rate:.1f}%",
                'color': 'p1'
            },
            {
                'label': 'Overall Progress',
                'value': f"{metrics.completion_rate:.1f}%",
                'color': None
            }
        ]

        row = self.add_metric_boxes(ws, kpi_metrics, start_row=row, boxes_per_row=4)

        # Summary statistics
        row += 2
        ws[f'A{row}'] = "Task Breakdown"
        ws.merge_cells(f'A{row}:F{row}')
        self.style_manager.apply_subheader_style(ws[f'A{row}'])

        row += 1
        summary_data = [
            ['Priority', 'Total', 'Completed', 'In Progress', 'Pending', 'Completion %'],
            [
                'P0 (Critical)',
                quarterly_data.p0_total,
                quarterly_data.p0_completed,
                quarterly_data.p0_in_progress,
                quarterly_data.p0_pending,
                f"{metrics.p0_completion_rate:.1f}%"
            ],
            [
                'P1 (High)',
                quarterly_data.p1_total,
                quarterly_data.p1_completed,
                quarterly_data.p1_in_progress,
                quarterly_data.p1_pending,
                f"{metrics.p1_completion_rate:.1f}%"
            ],
            [
                'P2 (Medium)',
                quarterly_data.p2_total,
                quarterly_data.p2_completed,
                quarterly_data.p2_in_progress,
                quarterly_data.p2_pending,
                f"{metrics.p2_completion_rate:.1f}%"
            ]
        ]

        row = self.add_table(ws, summary_data[0], summary_data[1:], start_row=row)

        # Apply priority color coding to rows
        for i, priority in enumerate(['P0', 'P1', 'P2'], start=row-2):
            self.style_manager.apply_priority_style(ws[f'A{i}'], priority)

        # Auto-adjust column widths
        self.style_manager.auto_adjust_column_widths(ws)

    def _create_sheet2_p0_tasks(self, quarterly_data):
        """
        Sheet 2: P0 Tasks (Critical)

        Table with all P0 tasks and their details
        """
        ws = self.add_sheet("P0 Tasks", 1)

        # Title
        ws['A1'] = "P0 Tasks - Critical Priority"
        ws.merge_cells('A1:H1')
        self.style_manager.apply_header_style(ws['A1'])
        ws.row_dimensions[1].height = 30

        # Check if we have P0 tasks
        if not quarterly_data.p0_tasks:
            ws['A3'] = "No P0 tasks found for this quarter"
            return

        # Table header
        row = 3
        headers = [
            'ID',
            'Description',
            'Status',
            'Priority',
            'Due Date',
            'Tags',
            'Progress',
            'Notes'
        ]

        # Prepare data
        task_data = []
        for idx, task in enumerate(quarterly_data.p0_tasks, start=1):
            # Determine status text
            if task.status:
                status_text = "✅ Completed"
            else:
                status_text = "🔄 In Progress"

            # Format due date
            due_date_str = task.due_date.strftime('%Y-%m-%d') if task.due_date else "No due date"

            # Tags
            tags_str = ', '.join(task.tags) if task.tags else ""

            task_data.append([
                f"P0-{idx:02d}",
                task.content,
                status_text,
                "P0",
                due_date_str,
                tags_str,
                "100%" if task.status else "In Progress",
                ""  # Notes column for manual entry
            ])

        # Add table
        row = self.add_table(ws, headers, task_data, start_row=row)

        # Apply P0 styling to priority column (column D)
        for i in range(4, 4 + len(task_data)):
            self.style_manager.apply_priority_style(ws[f'D{i}'], 'P0')

        # Apply status styling to status column (column C)
        for i in range(4, 4 + len(task_data)):
            cell = ws[f'C{i}']
            if "Completed" in str(cell.value):
                self.style_manager.apply_status_style(cell, 'completed')
            else:
                self.style_manager.apply_status_style(cell, 'in_progress')

        # Highlight overdue tasks
        from datetime import datetime
        today = datetime.now()
        for i, task in enumerate(quarterly_data.p0_tasks, start=4):
            if task.due_date and task.due_date < today and not task.status:
                # Overdue and incomplete - highlight entire row in red
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    cell = ws[f'{col}{i}']
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

        # Auto-adjust column widths
        self.style_manager.auto_adjust_column_widths(ws)

        # Add summary at bottom
        row += 2
        ws[f'A{row}'] = "Summary:"
        ws[f'B{row}'] = f"Total P0 Tasks: {quarterly_data.p0_total}"
        row += 1
        ws[f'B{row}'] = f"Completed: {quarterly_data.p0_completed}"
        row += 1
        ws[f'B{row}'] = f"In Progress: {quarterly_data.p0_in_progress}"
        row += 1
        ws[f'B{row}'] = f"Pending: {quarterly_data.p0_pending}"

    def _create_sheet3_p1_tasks(self, quarterly_data):
        """
        Sheet 3: P1 Tasks (High Priority)

        Same structure as P0 tasks but for P1
        """
        ws = self.add_sheet("P1 Tasks", 2)

        # Title
        ws['A1'] = "P1 Tasks - High Priority"
        ws.merge_cells('A1:H1')
        self.style_manager.apply_header_style(ws['A1'])
        ws.row_dimensions[1].height = 30

        # Check if we have P1 tasks
        if not quarterly_data.p1_tasks:
            ws['A3'] = "No P1 tasks found for this quarter"
            return

        # Table header
        row = 3
        headers = [
            'ID',
            'Description',
            'Status',
            'Priority',
            'Due Date',
            'Tags',
            'Progress',
            'Notes'
        ]

        # Prepare data
        task_data = []
        for idx, task in enumerate(quarterly_data.p1_tasks, start=1):
            # Determine status text
            if task.status:
                status_text = "✅ Completed"
            else:
                status_text = "🔄 In Progress"

            # Format due date
            due_date_str = task.due_date.strftime('%Y-%m-%d') if task.due_date else "No due date"

            # Tags
            tags_str = ', '.join(task.tags) if task.tags else ""

            task_data.append([
                f"P1-{idx:02d}",
                task.content,
                status_text,
                "P1",
                due_date_str,
                tags_str,
                "100%" if task.status else "In Progress",
                ""  # Notes column
            ])

        # Add table
        row = self.add_table(ws, headers, task_data, start_row=row)

        # Apply P1 styling to priority column (column D)
        for i in range(4, 4 + len(task_data)):
            self.style_manager.apply_priority_style(ws[f'D{i}'], 'P1')

        # Apply status styling to status column (column C)
        for i in range(4, 4 + len(task_data)):
            cell = ws[f'C{i}']
            if "Completed" in str(cell.value):
                self.style_manager.apply_status_style(cell, 'completed')
            else:
                self.style_manager.apply_status_style(cell, 'in_progress')

        # Auto-adjust column widths
        self.style_manager.auto_adjust_column_widths(ws)

        # Add summary at bottom
        row += 2
        ws[f'A{row}'] = "Summary:"
        ws[f'B{row}'] = f"Total P1 Tasks: {quarterly_data.p1_total}"
        row += 1
        ws[f'B{row}'] = f"Completed: {quarterly_data.p1_completed}"
        row += 1
        ws[f'B{row}'] = f"In Progress: {quarterly_data.p1_in_progress}"
        row += 1
        ws[f'B{row}'] = f"Pending: {quarterly_data.p1_pending}"

    def _create_sheet4_progress_charts(self, quarterly_data, metrics):
        """
        Sheet 4: Progress Charts

        Visual analytics with charts
        Note: Chart generation will be added in Phase 2
        For now, includes summary data that can be charted
        """
        ws = self.add_sheet("Progress Charts", 3)

        # Title
        ws['A1'] = "Progress Analytics"
        ws.merge_cells('A1:F1')
        self.style_manager.apply_header_style(ws['A1'])
        ws.row_dimensions[1].height = 30

        # Completion summary for pie chart
        row = 3
        ws[f'A{row}'] = "Task Completion Summary"
        ws.merge_cells(f'A{row}:C{row}')
        self.style_manager.apply_subheader_style(ws[f'A{row}'])

        row += 1
        completion_headers = ['Category', 'Count', 'Percentage']
        completion_data = [
            [
                'Completed',
                quarterly_data.p0_completed + quarterly_data.p1_completed + quarterly_data.p2_completed,
                f"{metrics.completion_rate:.1f}%"
            ],
            [
                'In Progress',
                quarterly_data.p0_in_progress + quarterly_data.p1_in_progress + quarterly_data.p2_in_progress,
                f"{(quarterly_data.p0_in_progress + quarterly_data.p1_in_progress + quarterly_data.p2_in_progress) / max(quarterly_data.total_tasks, 1) * 100:.1f}%"
            ],
            [
                'Pending',
                quarterly_data.p0_pending + quarterly_data.p1_pending + quarterly_data.p2_pending,
                f"{(quarterly_data.p0_pending + quarterly_data.p1_pending + quarterly_data.p2_pending) / max(quarterly_data.total_tasks, 1) * 100:.1f}%"
            ]
        ]

        row = self.add_table(ws, completion_headers, completion_data, start_row=row)

        # Priority breakdown
        row += 2
        ws[f'A{row}'] = "Completion by Priority"
        ws.merge_cells(f'A{row}:D{row}')
        self.style_manager.apply_subheader_style(ws[f'A{row}'])

        row += 1
        priority_headers = ['Priority', 'Total', 'Completed', 'Completion %']
        priority_data = [
            ['P0', quarterly_data.p0_total, quarterly_data.p0_completed, f"{metrics.p0_completion_rate:.1f}%"],
            ['P1', quarterly_data.p1_total, quarterly_data.p1_completed, f"{metrics.p1_completion_rate:.1f}%"],
            ['P2', quarterly_data.p2_total, quarterly_data.p2_completed, f"{metrics.p2_completion_rate:.1f}%"]
        ]

        row = self.add_table(ws, priority_headers, priority_data, start_row=row)

        # Apply priority styling
        for i, priority in enumerate(['P0', 'P1', 'P2'], start=row-2):
            self.style_manager.apply_priority_style(ws[f'A{i}'], priority)

        # Note about charts
        row += 3
        ws[f'A{row}'] = "📊 Note: Visual charts will be added in Phase 2 (Chart Integration)"
        ws.merge_cells(f'A{row}:F{row}')
        from openpyxl.styles import Font
        ws[f'A{row}'].font = Font(italic=True, size=10, color="808080")

        row += 1
        ws[f'A{row}'] = "Planned charts:"
        row += 1
        ws[f'B{row}'] = "• Completion pie chart (Completed vs Pending)"
        row += 1
        ws[f'B{row}'] = "• Priority breakdown bar chart"
        row += 1
        ws[f'B{row}'] = "• Weekly progress trend line"

        # Auto-adjust column widths
        self.style_manager.auto_adjust_column_widths(ws)
